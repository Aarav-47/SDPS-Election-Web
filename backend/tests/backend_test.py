"""SDPS Election Backend API tests - iteration 2 (users/posts/settings/reset)."""
import os
import io
import pytest
import requests
import openpyxl

BASE_URL = os.environ["REACT_APP_BACKEND_URL"].rstrip("/")
API = f"{BASE_URL}/api"


@pytest.fixture(scope="module")
def admin_token():
    r = requests.post(f"{API}/admin/login", json={"username": "Aarav", "password": "Krish@2026"})
    assert r.status_code == 200, r.text
    tok = r.json()["token"]
    assert isinstance(tok, str) and len(tok) > 10
    return tok


@pytest.fixture(scope="module")
def auth_headers(admin_token):
    return {"Authorization": f"Bearer {admin_token}"}


# ---------- Cleanup BEFORE tests start to make state deterministic ----------
@pytest.fixture(scope="module", autouse=True)
def reset_state(auth_headers):
    # Reset votes only - keep seed users/candidates
    requests.post(f"{API}/admin/reset/votes", headers=auth_headers)
    yield


# ---------- Public: Posts / Users / Candidates ----------
def test_root():
    r = requests.get(f"{API}/")
    assert r.status_code == 200


def test_posts_default_5_ordered():
    r = requests.get(f"{API}/posts")
    assert r.status_code == 200
    data = r.json()
    assert len(data) >= 5
    keys = [d["key"] for d in data[:5]]
    assert keys == ["head_boy", "head_girl", "sports_skipper", "cultural_head", "discipline_head"]
    # ordered by 'order' ascending
    orders = [d["order"] for d in data[:5]]
    assert orders == sorted(orders)


def test_get_user_student():
    r = requests.get(f"{API}/users/SDPSS001")
    assert r.status_code == 200
    d = r.json()
    assert d["role"] == "student"
    assert d["name"] == "Aarav Sharma"
    assert d["father_name"] == "Rajesh Sharma"
    assert d["class_name"] == "XII-A"
    assert "has_voted" in d


def test_get_user_teacher():
    r = requests.get(f"{API}/users/SDPSE01")
    assert r.status_code == 200
    d = r.json()
    assert d["role"] == "teacher"
    assert d["subject"] == "Mathematics"
    assert d["designation"] == "Sr. Teacher"


def test_get_user_404():
    r = requests.get(f"{API}/users/INVALID")
    assert r.status_code == 404


def test_public_settings_endpoint():
    r = requests.get(f"{API}/settings")
    assert r.status_code == 200
    assert isinstance(r.json(), dict)


# ---------- Voting ----------
def _selections():
    sel = {}
    for p in ["head_boy", "head_girl", "sports_skipper", "cultural_head", "discipline_head"]:
        r = requests.get(f"{API}/candidates", params={"post": p})
        assert r.status_code == 200
        sel[p] = r.json()[0]["id"]
    return sel


def test_vote_student_success():
    sel = _selections()
    r = requests.post(f"{API}/votes", json={"admission_no": "SDPSS003", "selections": sel})
    assert r.status_code == 200, r.text
    assert r.json()["ok"] is True
    # verify has_voted persisted
    g = requests.get(f"{API}/users/SDPSS003")
    assert g.json()["has_voted"] is True


def test_vote_teacher_success():
    sel = _selections()
    r = requests.post(f"{API}/votes", json={"admission_no": "SDPSE02", "selections": sel})
    assert r.status_code == 200, r.text


def test_vote_duplicate_rejected():
    sel = _selections()
    r = requests.post(f"{API}/votes", json={"admission_no": "SDPSS003", "selections": sel})
    assert r.status_code == 400
    assert "already" in r.text.lower()


# ---------- Admin: Posts (Categories) CRUD ----------
def test_admin_posts_crud(auth_headers):
    # Create
    r = requests.post(f"{API}/admin/posts", json={"title": "TEST_Vice Captain"}, headers=auth_headers)
    assert r.status_code == 200, r.text
    p = r.json()
    pid = p["id"]
    assert p["title"] == "TEST_Vice Captain"
    assert p["key"].startswith("test_vice_captain")

    # Update title
    r2 = requests.put(f"{API}/admin/posts/{pid}", json={"title": "TEST_Vice Captain Updated"}, headers=auth_headers)
    assert r2.status_code == 200
    assert r2.json()["title"] == "TEST_Vice Captain Updated"

    # GET /api/posts should reflect
    r3 = requests.get(f"{API}/posts")
    assert any(x["id"] == pid and x["title"] == "TEST_Vice Captain Updated" for x in r3.json())

    # Delete (no votes for it) -> success
    r4 = requests.delete(f"{API}/admin/posts/{pid}", headers=auth_headers)
    assert r4.status_code == 200


def test_admin_posts_delete_with_votes_blocked(auth_headers):
    # head_boy has votes (cast above); deleting must 400
    posts = requests.get(f"{API}/posts").json()
    hb = next(p for p in posts if p["key"] == "head_boy")
    r = requests.delete(f"{API}/admin/posts/{hb['id']}", headers=auth_headers)
    assert r.status_code == 400
    assert "votes" in r.text.lower()


# ---------- Admin: Users upload ----------
def _build_xlsx(rows):
    wb = openpyxl.Workbook()
    ws = wb.active
    for r in rows:
        ws.append(r)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def test_users_upload_student(auth_headers):
    content = _build_xlsx([
        ["admission_no", "name", "father_name", "class_name"],
        ["TEST_S100", "TEST Stud A", "TEST Father A", "XII-Z"],
        ["TEST_S101", "TEST Stud B", "TEST Father B", "XII-Z"],
    ])
    files = {"file": ("students.xlsx", content,
                      "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
    r = requests.post(f"{API}/admin/users/upload?role=student", files=files, headers=auth_headers)
    assert r.status_code == 200, r.text
    d = r.json()
    assert d["inserted"] + d["updated"] >= 2
    g = requests.get(f"{API}/users/TEST_S100").json()
    assert g["name"] == "TEST Stud A" and g["father_name"] == "TEST Father A" and g["role"] == "student"
    # cleanup
    requests.delete(f"{API}/admin/users/TEST_S100", headers=auth_headers)
    requests.delete(f"{API}/admin/users/TEST_S101", headers=auth_headers)


def test_users_upload_teacher(auth_headers):
    content = _build_xlsx([
        ["admission_no", "name", "subject", "designation"],
        ["TEST_T100", "TEST Teacher A", "Chemistry", "HOD"],
    ])
    files = {"file": ("teachers.xlsx", content,
                      "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
    r = requests.post(f"{API}/admin/users/upload?role=teacher", files=files, headers=auth_headers)
    assert r.status_code == 200, r.text
    g = requests.get(f"{API}/users/TEST_T100").json()
    assert g["role"] == "teacher" and g["subject"] == "Chemistry" and g["designation"] == "HOD"
    requests.delete(f"{API}/admin/users/TEST_T100", headers=auth_headers)


# ---------- Admin: Templates ----------
def test_template_student(auth_headers):
    r = requests.get(f"{API}/admin/template/student", headers=auth_headers)
    assert r.status_code == 200
    assert "spreadsheetml" in r.headers.get("content-type", "")
    wb = openpyxl.load_workbook(io.BytesIO(r.content))
    headers = [c.value for c in wb.active[1]]
    assert headers[:4] == ["admission_no", "name", "father_name", "class_name"]


def test_template_teacher(auth_headers):
    r = requests.get(f"{API}/admin/template/teacher", headers=auth_headers)
    assert r.status_code == 200
    assert "spreadsheetml" in r.headers.get("content-type", "")
    wb = openpyxl.load_workbook(io.BytesIO(r.content))
    headers = [c.value for c in wb.active[1]]
    assert headers[:4] == ["admission_no", "name", "subject", "designation"]


# ---------- Admin: Settings (school_logo) ----------
def test_settings_school_logo_persist(auth_headers):
    url_val = "https://example.com/test_logo.png"
    r = requests.put(f"{API}/admin/settings/school_logo",
                     json={"value": url_val}, headers=auth_headers)
    assert r.status_code == 200
    pub = requests.get(f"{API}/settings").json()
    assert pub.get("school_logo") == url_val


# ---------- Admin: Reset ----------
def test_reset_votes(auth_headers):
    r = requests.post(f"{API}/admin/reset/votes", headers=auth_headers)
    assert r.status_code == 200
    s = requests.get(f"{API}/admin/stats", headers=auth_headers).json()
    assert s["total_voted"] == 0


def test_reset_all_keeps_posts(auth_headers):
    r = requests.post(f"{API}/admin/reset/all", headers=auth_headers)
    assert r.status_code == 200
    s = requests.get(f"{API}/admin/stats", headers=auth_headers).json()
    assert s["total_voted"] == 0
    assert s["total_users"] == 0
    posts = requests.get(f"{API}/posts").json()
    assert len(posts) >= 5
    keys = [p["key"] for p in posts[:5]]
    assert keys == ["head_boy", "head_girl", "sports_skipper", "cultural_head", "discipline_head"]


# ---------- Auth negative ----------
def test_admin_login_wrong():
    r = requests.post(f"{API}/admin/login", json={"username": "Aarav", "password": "wrong"})
    assert r.status_code == 401


def test_admin_endpoint_no_token():
    r = requests.get(f"{API}/admin/users")
    assert r.status_code == 401
